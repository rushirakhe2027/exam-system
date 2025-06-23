from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort, send_file
from flask_login import login_required, current_user
from functools import wraps
from app.models.user import User
from app.models.class_model import Class
from app.models.submission import Submission
from app.mongodb import MongoManager, mongo
from datetime import datetime, timedelta
import secrets
import string
from bson import ObjectId
import json
from werkzeug.utils import secure_filename
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

teacher_bp = Blueprint('teacher', __name__, url_prefix='/teacher')

def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_teacher():
            return render_template('errors/403.html'), 403
        return f(*args, **kwargs)
    return decorated_function

def generate_password(length=10):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for i in range(length))

@teacher_bp.route('/dashboard')
@login_required
@teacher_required
def dashboard():
    # Get statistics
    total_students = MongoManager.count_students()
    total_exams = MongoManager.count_teacher_exams(current_user.id)
    active_exams_count = MongoManager.count_active_teacher_exams(current_user.id)
    recent_submissions = MongoManager.get_recent_submissions_for_teacher(current_user.id, limit=5)
    
    # Get recent exams (limit to 5, sorted by creation date)
    recent_exams = MongoManager.get_teacher_exams(current_user.id, limit=5, sort_by='created_at', sort_order=-1)
    
    # Get active exams list
    active_exams = MongoManager.get_active_teacher_exams(current_user.id, limit=5)
    
    # Calculate completion rate
    completion_rate = 0
    if total_exams > 0:
        completed_exams = MongoManager.count_completed_teacher_exams(current_user.id)
        completion_rate = round((completed_exams / total_exams) * 100)
    
    # Get pending reviews count
    pending_reviews = MongoManager.count_pending_reviews(current_user.id)
    
    return render_template('teacher/dashboard.html',
                         total_students=total_students,
                         total_exams=total_exams,
                         active_exams_count=active_exams_count,
                         active_exams=active_exams,
                         recent_submissions=recent_submissions,
                         recent_exams=recent_exams,
                         completion_rate=completion_rate,
                         pending_reviews=pending_reviews)

@teacher_bp.route('/exams')
@login_required
@teacher_required
def exams():
    teacher_exams = MongoManager.get_teacher_exams(current_user.id)
    return render_template('teacher/exams.html', exams=teacher_exams)

@teacher_bp.route('/exam/<exam_id>/details')
@login_required
@teacher_required
def exam_details(exam_id):
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam:
        flash('Exam not found', 'error')
        return redirect(url_for('teacher.dashboard'))

    # Get exam statistics
    total_submissions = MongoManager.count_exam_submissions(exam_id)
    completed_submissions = MongoManager.count_completed_submissions(exam_id)
    
    # Use the MongoManager method for average score calculation
    # This already handles both subjective (score = percentage) and objective exams correctly
    average_score = MongoManager.get_exam_average_score(exam_id)
    
    # Get recent submissions with enhanced data
    recent_submissions_raw = MongoManager.get_recent_submissions_for_exam(exam_id, limit=10)
    recent_submissions = []
    
    for sub_data in recent_submissions_raw:
        # Convert to submission object if it's raw data
        if isinstance(sub_data, dict):
            submission = Submission.from_db(sub_data)
        else:
            submission = sub_data
            
        # Add student name if not already present
        if not hasattr(submission, 'student_name') or not submission.student_name:
            student = MongoManager.get_user_by_id(submission.student_id)
            submission.student_name = student.username if student else 'Unknown'
        
        # Calculate proper score display based on exam format
        if submission.is_graded and submission.score is not None:
            if exam.exam_format == 'subjective':
                # For subjective exams, score is already percentage
                percentage = submission.score
                marks_earned = (percentage * exam.total_marks / 100) if exam.total_marks else 0
                submission.marks_display = f"{marks_earned:.1f}/{exam.total_marks or 0}"
                submission.percentage_display = f"{percentage:.1f}%"
            elif hasattr(submission, 'detailed_results') and submission.detailed_results:
                # Use detailed results for accurate marks calculation (objective exams)
                total_marks_earned = sum(result.get('score', 0) for result in submission.detailed_results)
                total_possible_marks = sum(result.get('marks', 0) for result in submission.detailed_results)
                submission.marks_display = f"{total_marks_earned}/{total_possible_marks}"
                submission.percentage_display = f"{(total_marks_earned / total_possible_marks * 100):.1f}%" if total_possible_marks > 0 else "0.0%"
            elif hasattr(submission, 'percentage') and submission.percentage is not None:
                # Use stored percentage
                marks_earned = (submission.percentage * exam.total_marks / 100) if exam.total_marks else submission.score
                submission.marks_display = f"{marks_earned:.0f}/{exam.total_marks or 100}"
                submission.percentage_display = f"{submission.percentage:.1f}%"
            else:
                # Calculate from score and total marks
                if exam.total_marks and exam.total_marks > 0:
                    percentage = (submission.score / exam.total_marks) * 100
                    submission.marks_display = f"{submission.score:.0f}/{exam.total_marks}"
                    submission.percentage_display = f"{percentage:.1f}%"
                else:
                    submission.marks_display = f"{submission.score:.0f}/100"
                    submission.percentage_display = f"{submission.score:.1f}%"
        else:
            submission.marks_display = "Pending"
            submission.percentage_display = "Pending"
        
        recent_submissions.append(submission)

    # Add total questions count and questions attended
    exam.total_questions = len(exam.questions) if hasattr(exam, 'questions') else 0
    
    # Calculate questions attended by students
    questions_attended = 0
    if recent_submissions:
        # Get the maximum number of questions any student answered
        for submission in recent_submissions:
            if hasattr(submission, 'answers') and submission.answers:
                if isinstance(submission.answers, dict):
                    questions_attended = max(questions_attended, len(submission.answers))
                elif isinstance(submission.answers, list):
                    questions_attended = max(questions_attended, len(submission.answers))
    
    # If no submissions yet, use total questions
    if questions_attended == 0:
        questions_attended = exam.total_questions

    # Get class information
    class_name = "Not assigned to any class"
    if hasattr(exam, 'class_id') and exam.class_id:
        class_obj = MongoManager.get_class_by_id(exam.class_id)
        if class_obj:
            class_name = class_obj.name

    # Fetch all classes for the current teacher
    classes = MongoManager.get_teacher_classes(current_user.id)

    return render_template('teacher/exam_details.html',
                         exam=exam,
                         total_submissions=total_submissions,
                         completed_submissions=completed_submissions,
                         average_score=average_score,
                         recent_submissions=recent_submissions,
                         questions_attended=questions_attended,
                         class_name=class_name,
                         classes=classes)

@teacher_bp.route('/exam/create', methods=['GET', 'POST'])
@login_required
@teacher_required
def create_exam():
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # Get all classes for the current teacher
    classes = MongoManager.get_teacher_classes(current_user.id)
    
    if request.method == 'POST':
        # Skip form validation since CSRF is disabled
        # if not form.validate():
        #     flash('Form validation failed. Please try again.', 'error')
        #     return render_template('teacher/create_exam.html', form=form, classes=classes)
            
        # Get form data
        title = request.form.get('title')
        description = request.form.get('description')
        duration = request.form.get('duration')
        total_marks = request.form.get('total_marks')
        exam_type = 'self_paced'  # Force all exams to be self_paced
        exam_format = request.form.get('exam_format')
        end_time = request.form.get('end_time')
        num_questions_to_display = request.form.get('num_questions_to_display')
        class_id = request.form.get('class_id', '')  # Get class_id
        
        # Validate required fields
        if not all([title, duration, total_marks, exam_format, end_time]):
            flash('Please fill in all required fields.', 'error')
            return render_template('teacher/create_exam.html', classes=classes)
            
        # Validate exam format
        if exam_format not in ['objective', 'subjective']:
            flash('Invalid exam format.', 'error')
            return render_template('teacher/create_exam.html', classes=classes)
            
        # Convert string dates to datetime
        try:
            # Handle end_time
            if not end_time:
                return jsonify({'error': 'Due date is required'}), 400
            
            # Parse end_time
            try:
                # First try direct ISO format
                end_time = datetime.fromisoformat(end_time)
            except ValueError:
                # If that fails, try parsing the date-time picker format
                end_time = datetime.strptime(end_time, '%Y-%m-%dT%H:%M')
            
            # Self-paced exams don't have start time
            start_time = None
            
            # Ensure end time is in the future
            if end_time <= datetime.utcnow():
                return jsonify({'error': 'Due date must be in the future'}), 400
                
            print(f"Parsed dates - end: {end_time}")
            
        except ValueError as e:
            print(f"Error parsing dates: {str(e)}")
            return jsonify({'error': 'Invalid date format. Please use format: YYYY-MM-DD HH:MM'}), 400
            
        # Convert duration and total_marks to integers
        try:
            duration = int(duration)
            total_marks = int(total_marks)
            
            if duration < 1 or duration > 480:
                flash('Duration must be between 1 and 480 minutes.', 'error')
                return render_template('teacher/create_exam.html', classes=classes)
                
            if total_marks < 1 or total_marks > 1000:
                flash('Total marks must be between 1 and 1000.', 'error')
                return render_template('teacher/create_exam.html', classes=classes)
                
            # Convert num_questions_to_display to integer if provided
            if num_questions_to_display:
                num_questions_to_display = int(num_questions_to_display)
                if num_questions_to_display < 1:
                    flash('Number of questions to display must be at least 1.', 'error')
                    return render_template('teacher/create_exam.html', classes=classes)
            else:
                num_questions_to_display = None
                
        except ValueError:
            flash('Invalid numeric values provided.', 'error')
            return render_template('teacher/create_exam.html', classes=classes)
            
        # Convert class_id to ObjectId if provided
        if class_id:
            try:
                class_id = ObjectId(class_id)
            except:
                flash('Invalid class ID format.', 'error')
                return render_template('teacher/create_exam.html', classes=classes)
        else:
            class_id = None
            
        # Create the exam
        exam = MongoManager.create_exam(
            title=title,
            description=description,
            duration=duration,
            teacher_id=current_user.id,
            start_time=start_time,
            end_time=end_time,
            total_marks=total_marks,
            exam_type=exam_type,
            exam_format=exam_format,
            num_questions_to_display=num_questions_to_display,
            class_id=class_id  # Add class_id to exam creation
        )
        
        if exam:
            flash('Exam created successfully!', 'success')
            return redirect(url_for('teacher.manage_questions', exam_id=exam.id))
        else:
            flash('Failed to create exam. Please try again.', 'error')
            
    return render_template('teacher/create_exam.html', classes=classes)

@teacher_bp.route('/exam/<exam_id>/edit', methods=['GET'])
@login_required
@teacher_required
def edit_exam(exam_id):
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam or str(exam.teacher_id) != str(current_user.id):
        abort(404)
    
    # Fetch all classes for the current teacher
    classes = MongoManager.get_teacher_classes(current_user.id)
    
    return render_template('teacher/edit_exam.html', exam=exam, classes=classes)

@teacher_bp.route('/exam/<exam_id>/delete', methods=['POST'])
@login_required
@teacher_required
def delete_exam(exam_id):
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam or exam.teacher_id != current_user.id:
        abort(404)
        
    MongoManager.delete_exam(exam_id)
    flash('Exam deleted successfully!', 'success')
    return redirect(url_for('teacher.exams'))

@teacher_bp.route('/exam/<exam_id>/questions', methods=['GET', 'POST'])
@login_required
@teacher_required
def manage_questions(exam_id):
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam or str(exam.teacher_id) != str(current_user.id):
        abort(403)
    
    if request.method == 'POST':
        # Skip form validation since CSRF is disabled
        # if not form.validate():
        #     flash('Form validation failed. Please try again.', 'error')
        #     return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
        try:
            # Get basic question data
            question_type = request.form.get('question_type')
            question_text = request.form.get('question_text')
            marks = int(request.form.get('marks', 1))
            
            # Validate required fields
            if not all([question_type, question_text]):
                flash('Please fill in all required fields.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            # Validate question type
            if question_type not in ['multiple_choice', 'subjective']:
                flash('Invalid question type.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            # Prepare question data
            question_data = {
                'question_type': question_type,
                'question_text': question_text,
                'marks': marks
            }
            
            # Handle multiple choice questions
            if question_type == 'multiple_choice':
                options = request.form.getlist('options[]')
                correct_answer = request.form.get('correct_answer')
                
                # Validate multiple choice data
                if not options or not correct_answer:
                    flash('Please provide all options and select a correct answer for multiple choice questions.', 'error')
                    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
                
                # Clean and validate options
                options = [opt.strip() for opt in options if opt.strip()]
                if len(options) < 2:
                    flash('Multiple choice questions must have at least 2 options.', 'error')
                    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
                
                try:
                    correct_answer_index = int(correct_answer)
                    if not (0 <= correct_answer_index < len(options)):
                        flash('Invalid correct answer selection.', 'error')
                        return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
                    
                    # Store the actual answer text instead of index
                    correct_answer_text = options[correct_answer_index].strip()
                    
                except ValueError:
                    flash('Invalid correct answer value.', 'error')
                    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
                
                question_data.update({
                    'options': options,
                    'correct_answer': correct_answer_text  # Store actual text, not index
                })
            else:  # descriptive
                model_answer = request.form.get('model_answer', '').strip()
                question_data['model_answer'] = model_answer
            
            # Add the question to the exam
            if MongoManager.add_question(exam_id, question_data):
                flash('Question added successfully!', 'success')
            else:
                flash('Failed to add question. Please try again.', 'error')
                
        except Exception as e:
            print(f"Error adding question: {str(e)}")
            flash('An error occurred while adding the question.', 'error')
            
        return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
    
    return render_template('teacher/manage_questions.html', exam=exam)

@teacher_bp.route('/exam/<exam_id>/question/<question_id>/update', methods=['POST'])
@login_required
@teacher_required
def update_question(exam_id, question_id):
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     flash('Form validation failed. Please try again.', 'error')
    #     return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam or str(exam.teacher_id) != str(current_user.id):
        abort(404)
    
    try:
        # Get basic question data
        question_text = request.form.get('question_text')
        marks = int(request.form.get('marks', 1))
        
        # Get the existing question to determine its type
        existing_question = next((q for q in exam.questions if str(q.id) == question_id), None)
        if not existing_question:
            flash('Question not found.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        # Prepare update data
        update_data = {
            'question_text': question_text,
            'marks': marks,
            'question_type': existing_question.question_type
        }
        
        # Handle type-specific updates
        if existing_question.question_type == 'multiple_choice':
            options = request.form.getlist('options[]')
            correct_answer = request.form.get('correct_answer')
            
            # Validate multiple choice data
            if not options or not correct_answer:
                flash('Please provide all options and select a correct answer for multiple choice questions.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            # Clean and validate options
            options = [opt.strip() for opt in options if opt.strip()]
            if len(options) < 2:
                flash('Multiple choice questions must have at least 2 options.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            try:
                correct_answer_index = int(correct_answer)
                if not (0 <= correct_answer_index < len(options)):
                    flash('Invalid correct answer selection.', 'error')
                    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
                
                # Store the actual answer text instead of index
                correct_answer_text = options[correct_answer_index].strip()
                
            except ValueError:
                flash('Invalid correct answer value.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            update_data.update({
                'options': options,
                'correct_answer': correct_answer_text  # Store actual text, not index
            })
        else:  # descriptive
            model_answer = request.form.get('model_answer', '').strip()
            update_data['model_answer'] = model_answer
        
        # Update the question
        if MongoManager.update_question(exam_id, question_id, update_data):
            flash('Question updated successfully!', 'success')
        else:
            flash('Failed to update question.', 'error')
            
    except (ValueError, TypeError) as e:
        print(f"Error updating question: {str(e)}")
        flash('Invalid input. Please check your values.', 'error')
        
    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))

@teacher_bp.route('/exam/<exam_id>/question/<question_id>/delete', methods=['POST'])
@login_required
@teacher_required
def delete_question(exam_id, question_id):
    print(f"Delete request received - exam_id: {exam_id}, question_id: {question_id}")
    
    try:
        # Get the exam
        exam = MongoManager.get_exam_by_id(exam_id)
        if not exam:
            print("Exam not found")
            flash('Exam not found.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        # Check ownership
        if str(exam.teacher_id) != str(current_user.id):
            print("Unauthorized access attempt")
            flash('Unauthorized access.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        # Find the question in the exam's questions
        question_exists = False
        for question in exam.questions:
            if str(question.id) == str(question_id):
                question_exists = True
                break
        
        if not question_exists:
            print("Question not found in exam")
            flash('Question not found in exam.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        # Delete the question
        if MongoManager.delete_question(exam_id, question_id):
            print("Question deleted successfully")
            flash('Question deleted successfully!', 'success')
        else:
            print("Failed to delete question")
            flash('Failed to delete question.', 'error')
            
    except Exception as e:
        print(f"Error in delete_question route: {str(e)}")
        flash('An error occurred while deleting the question.', 'error')
    
    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))

@teacher_bp.route('/students')
@login_required
@teacher_required
def students():
    # Get all students using MongoManager
    students = MongoManager.get_all_students()
    # Get all classes for the current teacher
    classes = MongoManager.get_teacher_classes(current_user.id)
    return render_template('teacher/students.html', students=students, classes=classes)

@teacher_bp.route('/students/<student_id>')
@login_required
@teacher_required
def view_student(student_id):
    student = MongoManager.get_user_by_id(student_id)
    if not student or not student.is_student():
        return render_template('errors/403.html'), 403
    
    # Get student's class information
    class_data = mongo.db.classes.find_one({'students': ObjectId(student_id)})
    if class_data:
        student.class_info = {
            'id': str(class_data['_id']),
            'name': class_data['name']
        }
    
    # Verify student belongs to one of teacher's classes
    teacher_classes = MongoManager.get_teacher_classes(current_user.id)
    student_in_teacher_class = any(
        str(student.id) in [str(s_id) for s_id in class_obj.students]
        for class_obj in teacher_classes
    )
    
    if not student_in_teacher_class:
        return render_template('errors/403.html'), 403
    
    # Get student's submissions for this teacher's exams
    teacher_exam_ids = [exam.id for exam in MongoManager.find_exams(teacher_id=current_user.id)]
    submissions = MongoManager.find_submissions(student_id=student_id, exam_id=teacher_exam_ids)
    
    # Add exam information to each submission
    for submission in submissions:
        exam = MongoManager.get_exam_by_id(submission.exam_id)
        if exam:
            submission.exam_title = exam.title
        else:
            submission.exam_title = 'Unknown Exam'
    
    # Get all classes for the current teacher (for the edit modal)
    classes = MongoManager.get_teacher_classes(current_user.id)
    
    return render_template('teacher/student_details.html', 
                         student=student, 
                         submissions=submissions,
                         classes=classes)

@teacher_bp.route('/analytics')
@login_required
@teacher_required
def analytics():
    # Get analytics data
    analytics_data = MongoManager.get_exam_analytics(current_user.id)
    total_students = MongoManager.count_students()
    
    return render_template('teacher/analytics.html',
                         total_exams=analytics_data['total_exams'],
                         total_students=total_students,
                         total_submissions=analytics_data['total_submissions'],
                         avg_score=analytics_data['avg_score'])

@teacher_bp.route('/settings', methods=['GET', 'POST'])
@login_required
@teacher_required
def settings():
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    if request.method == 'POST':
        # Skip form validation since CSRF is disabled
        # if not form.validate():
        #     flash('Form validation failed. Please try again.', 'error')
        #     return render_template('teacher/settings.html', form=form)
            
        # Update profile information
        update_data = {
            'username': request.form.get('username'),
            'email': request.form.get('email')
        }
        
        # Update password if provided
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        if current_password and new_password:
            user = MongoManager.get_user_by_id(current_user.id)
            if user and user.check_password(current_password):
                update_data['password_hash'] = User.set_password(new_password)
            else:
                flash('Current password is incorrect', 'error')
                return redirect(url_for('teacher.settings'))
        
        if MongoManager.update_user(current_user.id, update_data):
            flash('Settings updated successfully!', 'success')
        else:
            flash('Failed to update settings', 'error')
        return redirect(url_for('teacher.settings'))
    
    return render_template('teacher/settings.html')

@teacher_bp.route('/class/<class_id>')
@login_required
@teacher_required
def view_class(class_id):
    class_obj = MongoManager.get_class_by_id(class_id)
    if not class_obj or str(class_obj.teacher_id) != str(current_user.id):
        return render_template('errors/403.html'), 403
    
    # Get students in this class
    class_students = MongoManager.get_class_students(class_id)
    class_obj.students = class_students
    
    # Get all students not in this class for the add student modal
    all_students = MongoManager.get_all_students()
    class_student_ids = [str(s.id) for s in class_students]
    available_students = [s for s in all_students if str(s.id) not in class_student_ids]
    
    return render_template('teacher/class_details.html', 
                         class_obj=class_obj, 
                         students=available_students)

@teacher_bp.route('/class/<class_id>/delete', methods=['POST'])
@login_required
@teacher_required
def delete_class(class_id):
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     flash('Form validation failed. Please try again.', 'error')
    #     return redirect(url_for('teacher.students'))
        
    class_obj = MongoManager.get_class_by_id(class_id)
    if not class_obj or str(class_obj.teacher_id) != str(current_user.id):
        flash('Permission denied', 'error')
        return redirect(url_for('teacher.students'))
    
    if MongoManager.delete_class(class_id):
        flash('Class deleted successfully', 'success')
    else:
        flash('Failed to delete class', 'error')
    
    return redirect(url_for('teacher.students'))

@teacher_bp.route('/class/create', methods=['POST'])
@login_required
@teacher_required
def create_class():
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     flash('Form validation failed. Please try again.', 'error')
    #     return redirect(url_for('teacher.students'))
        
    name = request.form.get('name')
    description = request.form.get('description')
    
    if not name:
        flash('Class name is required!', 'error')
        return redirect(url_for('teacher.students'))
    
    class_obj = Class(
        name=name,
        description=description,
        teacher_id=current_user.id
    )
    
    MongoManager.create_class(class_obj)
    flash('Class created successfully!', 'success')
    return redirect(url_for('teacher.students'))

@teacher_bp.route('/class/<class_id>/add_student', methods=['POST'])
@login_required
@teacher_required
def add_student_to_class(class_id):
    try:
        # Skip CSRF validation since it's disabled in the app
        # from flask_wtf import FlaskForm
        # form = FlaskForm()
        
        # if not form.validate():
        #     return jsonify({'error': 'CSRF validation failed'}), 400
        
        class_obj = MongoManager.get_class_by_id(class_id)
        if not class_obj or str(class_obj.teacher_id) != str(current_user.id):
            return jsonify({'error': 'Permission denied'}), 403
        
        student_id = request.form.get('student_id')
        if not student_id:
            return jsonify({'error': 'Please select a student'}), 400
        
        student = MongoManager.get_user_by_id(student_id)
        if not student or not student.is_student():
            return jsonify({'error': 'Invalid student selected'}), 400
            
        # Check if student is already in any class
        current_class = MongoManager.get_student_class(student_id)
        if current_class:
            return jsonify({'error': f'Student is already enrolled in class: {current_class.name}'}), 400
            
        # Check if student is already in this class
        if str(student.id) in [str(s_id) for s_id in class_obj.students]:
            return jsonify({'error': 'Student is already in this class'}), 400
        
        # Add student to class
        if MongoManager.add_student_to_class(class_id, student_id):
            return jsonify({'success': True, 'message': 'Student added successfully'})
            
    except Exception as e:
        print(f"Error adding student to class: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@teacher_bp.route('/class/<class_id>/remove_student/<student_id>', methods=['POST'])
@login_required
@teacher_required
def remove_student_from_class(class_id, student_id):
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     return jsonify({'success': False, 'error': 'CSRF validation failed'}), 400
        
    class_obj = MongoManager.get_class_by_id(class_id)
    if not class_obj or str(class_obj.teacher_id) != str(current_user.id):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    if str(student_id) not in [str(s_id) for s_id in class_obj.students]:
        return jsonify({'error': 'Student not in class'}), 400
    
    if MongoManager.remove_student_from_class(class_id, student_id):
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to remove student from class'}), 500

@teacher_bp.route('/student/add', methods=['POST'])
@login_required
@teacher_required
def add_student():
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     flash('Form validation failed. Please try again.', 'error')
    #     return redirect(url_for('teacher.students'))
        
    # Get form data
    username = request.form.get('username')
    email = request.form.get('email')
    student_id = request.form.get('student_id')
    phone_number = request.form.get('phone_number')
    college_name = request.form.get('college_name')
    class_id = request.form.get('class_id')
    
    # Generate a random password if not provided
    password = request.form.get('password') or generate_password()
    
    # Validate required fields
    if not all([username, email, student_id]):
        flash('Username, email and student ID are required!', 'error')
        return redirect(url_for('teacher.students'))
    
    # Validate unique fields
    if MongoManager.get_user_by_username(username):
        flash('Username already exists!', 'error')
        return redirect(url_for('teacher.students'))
    
    if MongoManager.get_user_by_email(email):
        flash('Email already registered!', 'error')
        return redirect(url_for('teacher.students'))
    
    if MongoManager.get_user_by_student_id(student_id):
        flash('Student ID already exists!', 'error')
        return redirect(url_for('teacher.students'))
    
    try:
        # Create new student
        student_data = {
            'username': username,
            'email': email,
            'student_id': student_id,
            'phone_number': phone_number,
            'college_name': college_name,
            'password': password,
            'role': 'student'  # Ensure role is set
        }
        
        print(f"\n=== CREATING STUDENT ===")
        print(f"Student data: {student_data}")
        print(f"Class ID from form: {class_id}")
        
        student = MongoManager.create_student(student_data)
        
        if not student:
            flash('Failed to create student!', 'error')
            return redirect(url_for('teacher.students'))
        
        print(f"Student created with ID: {student.id}")
        
        # Add student to class if specified
        if class_id:
            print(f"Attempting to add student to class {class_id}")
            class_obj = MongoManager.get_class_by_id(class_id)
            print(f"Class object: {class_obj}")
            
            if class_obj:
                print(f"Class teacher ID: {class_obj.teacher_id}, Current user ID: {current_user.id}")
                print(f"Teacher match: {str(class_obj.teacher_id) == str(current_user.id)}")
                
                if str(class_obj.teacher_id) == str(current_user.id):
                    add_result = MongoManager.add_student_to_class(class_id, str(student.id))
                    print(f"Add to class result: {add_result}")
                    
                    if add_result:
                        flash(f'Student added successfully and assigned to class "{class_obj.name}"!', 'success')
                    else:
                        flash('Student created but could not be added to class!', 'warning')
                else:
                    flash('Student created but invalid class specified!', 'warning')
            else:
                flash('Student created but class not found!', 'warning')
        else:
            print("No class ID provided")
            flash('Student added successfully!', 'success')
        
        return redirect(url_for('teacher.students'))
        
    except Exception as e:
        flash(f'Error creating student: {str(e)}', 'error')
        return redirect(url_for('teacher.students'))

@teacher_bp.route('/students/<student_id>/delete', methods=['POST'])
@login_required
@teacher_required
def delete_student(student_id):
    try:
        student = MongoManager.get_user_by_id(ObjectId(student_id))
        if not student or not student.is_student():
            return jsonify({'error': 'Invalid student'}), 400
        
        # Check if student is in any of teacher's classes
        teacher_classes = MongoManager.get_teacher_classes(current_user.id)
        student_in_teacher_class = any(
            str(student._id) in [str(s_id) for s_id in class_obj.students]
            for class_obj in teacher_classes
        )
        
        if not student_in_teacher_class:
            return jsonify({'error': 'Unauthorized'}), 403
        
        if MongoManager.delete_student(student._id):  # Use _id directly
            return jsonify({'message': 'Student deleted successfully'})
        else:
            return jsonify({'error': 'Failed to delete student'}), 500
    except Exception as e:
        print(f"Error in delete_student: {str(e)}")  # Add logging
        return jsonify({'error': str(e)}), 400

@teacher_bp.route('/exam/<exam_id>/add-questions', methods=['GET', 'POST'])
@login_required
@teacher_required
def add_exam_questions(exam_id):
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam or str(exam.teacher_id) != str(current_user.id):
        abort(403)
    
    if request.method == 'POST':
        # Skip form validation since CSRF is disabled
        # if not form.validate():
        #     flash('Form validation failed. Please try again.', 'error')
        #     return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
        question_ids = request.form.getlist('question_ids[]')
        success_count = 0
        for question_id in question_ids:
            if MongoManager.add_question_from_bank(exam_id, question_id):
                success_count += 1
        
        if success_count > 0:
            flash(f'{success_count} questions added to exam successfully!', 'success')
        else:
            flash('No questions were added to the exam.', 'warning')
        return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
    
    # Get questions from bank for selection
    available_questions = MongoManager.get_teacher_question_bank(current_user.id)
    return render_template('teacher/manage_questions.html', 
                         exam=exam, 
                         available_questions=available_questions,
                         form=form)

@teacher_bp.route('/question-bank')
@login_required
@teacher_required
def question_bank():
    questions = MongoManager.get_teacher_question_bank(current_user.id)
    return render_template('teacher/question_bank.html', questions=questions)

@teacher_bp.route('/question-bank/add', methods=['GET', 'POST'])
@login_required
@teacher_required
def add_to_bank():
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    if request.method == 'POST':
        try:
            # Skip form validation since CSRF is disabled
            # if not form.validate():
            #     flash('Form validation failed. Please try again.', 'error')
            #     return redirect(url_for('teacher.add_to_bank'))
                
            # Debug logging
            print("\n=== Adding Question to Bank ===")
            print("Form data received:")
            for key, value in request.form.items():
                print(f"{key}: {value}")
            
            # Prepare question data with consistent field names
            question_data = {
                'text': request.form['question_text'],  # Store as 'text' in DB
                'type': request.form['question_type'],  # Store as 'type' in DB
                'points': int(request.form['marks']),   # Store as 'points' in DB
                'difficulty_level': request.form.get('difficulty_level', 'medium'),
                'subject': request.form.get('subject', ''),
                'topic': request.form.get('topic', '')
            }
            
            print("\nProcessed base question data:", question_data)
            
            if question_data['type'] == 'multiple_choice':
                options = request.form.getlist('options[]')
                correct_answer = request.form.get('correct_answer')
                
                print("\nMultiple choice data:")
                print(f"Options: {options}")
                print(f"Correct answer: {correct_answer}")
                
                # Validate multiple choice data
                if not options or not correct_answer:
                    flash('Please provide all options and select a correct answer for multiple choice questions.', 'error')
                    return redirect(url_for('teacher.add_to_bank'))
                
                try:
                    correct_answer_index = int(correct_answer)
                    if not (0 <= correct_answer_index < len(options)):
                        flash('Invalid correct answer selection.', 'error')
                        return redirect(url_for('teacher.add_to_bank'))
                    
                    # Store the actual answer text instead of index
                    correct_answer_text = options[correct_answer_index].strip()
                    
                except ValueError:
                    flash('Invalid correct answer value.', 'error')
                    return redirect(url_for('teacher.add_to_bank'))
                
                # Remove any empty options
                options = [opt.strip() for opt in options if opt.strip()]
                if len(options) < 2:
                    flash('Multiple choice questions must have at least 2 options.', 'error')
                    return redirect(url_for('teacher.add_to_bank'))
                
                question_data.update({
                    'options': options,
                    'correct_answer': correct_answer_text  # Store actual text, not index
                })
            else:  # subjective
                question_data['model_answer'] = request.form.get('model_answer', '')
            
            print("\nFinal question data being sent to DB:", question_data)
            
            # Add to database
            result = MongoManager.add_to_question_bank(question_data, current_user.id)
            if result:
                print(f"\nQuestion added successfully with ID: {result}")
                flash('Question added to bank successfully!', 'success')
            else:
                print("\nFailed to add question to bank")
                flash('Failed to add question to bank.', 'error')
                
        except Exception as e:
            print(f"\nError in add_to_bank: {str(e)}")
            flash('An error occurred while adding the question.', 'error')
            
        return redirect(url_for('teacher.question_bank'))
    
    return render_template('teacher/add_question.html', form=form)

@teacher_bp.route('/student/<student_id>', methods=['GET'])
@login_required
@teacher_required
def get_student(student_id):
    try:
        student = MongoManager.get_user_by_id(student_id)
        if not student or not student.is_student():
            return jsonify({'error': 'Student not found'}), 404
        
        # Get student's current class
        class_data = mongo.db.classes.find_one({'students': ObjectId(student_id)})
        class_info = {
            'id': str(class_data['_id']) if class_data else None,
            'name': class_data['name'] if class_data else None
        } if class_data else None
        
        # Convert ObjectId to string for JSON serialization
        student_data = {
            'id': str(student.id),
            'username': student.username,
            'email': student.email,
            'student_id': student.student_id,
            'phone_number': student.phone_number or '',
            'college_name': student.college_name or '',
            'class_info': class_info
        }
        
        return jsonify(student_data)
    except Exception as e:
        print(f"Error in get_student: {str(e)}")
        return jsonify({'error': str(e)}), 400

@teacher_bp.route('/student/update', methods=['POST'])
@login_required
@teacher_required
def update_student():
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     flash('Form validation failed. Please try again.', 'error')
    #     return redirect(url_for('teacher.students'))
        
    try:
        student_id = request.form.get('_id')
        student = MongoManager.get_user_by_id(student_id)
        
        if not student or not student.is_student():
            flash('Student not found', 'error')
            return redirect(url_for('teacher.students'))
        
        # Check if username or email is being changed and verify uniqueness
        new_username = request.form.get('username')
        new_email = request.form.get('email')
        new_student_id = request.form.get('student_id')
        
        if new_username != student.username:
            existing = MongoManager.get_user_by_username(new_username)
            if existing and str(existing.id) != str(student_id):
                flash('Username already taken', 'error')
                return redirect(url_for('teacher.students'))
        
        if new_email != student.email:
            existing = MongoManager.get_user_by_email(new_email)
            if existing and str(existing.id) != str(student_id):
                flash('Email already registered', 'error')
                return redirect(url_for('teacher.students'))
                
        if new_student_id != student.student_id:
            existing = MongoManager.get_user_by_student_id(new_student_id)
            if existing and str(existing.id) != str(student_id):
                flash('Student ID already exists', 'error')
                return redirect(url_for('teacher.students'))
        
        # Prepare update data
        update_data = {
            'username': new_username,
            'email': new_email,
            'student_id': new_student_id,
            'phone_number': request.form.get('phone_number'),
            'college_name': request.form.get('college_name')
        }
        
        # Update password if provided
        new_password = request.form.get('password')
        if new_password:
            update_data['password_hash'] = User.set_password(new_password)
        
        # Handle class assignment
        new_class_id = request.form.get('class_id')
        if new_class_id:
            # Remove student from all current classes
            MongoManager.remove_student_from_all_classes(student_id)
            # Add to new class
            MongoManager.add_student_to_class(new_class_id, student_id)
        
        # Update student
        if MongoManager.update_user(student_id, update_data):
            flash('Student updated successfully', 'success')
        else:
            flash('Failed to update student', 'error')
        
        return redirect(url_for('teacher.students'))
    except Exception as e:
        print(f"Error in update_student: {str(e)}")
        flash(f'Error updating student: {str(e)}', 'error')
        return redirect(url_for('teacher.students'))

@teacher_bp.route('/submission/<submission_id>')
@login_required
@teacher_required
def view_submission(submission_id):
    # Get the submission
    submission = MongoManager.get_submission_by_id(submission_id)
    if not submission:
        return render_template('errors/404.html'), 404
    
    # Get the exam
    exam = MongoManager.get_exam_by_id(submission.exam_id)
    if not exam:
        return render_template('errors/404.html'), 404
    
    # Verify the exam belongs to the current teacher
    if str(exam.teacher_id) != str(current_user.id):
        return render_template('errors/403.html'), 403
    
    # Get the student
    student = MongoManager.get_user_by_id(submission.student_id)
    if not student:
        return render_template('errors/404.html'), 404
    
    # Process submission answers to match with exam questions
    processed_answers = []
    
    print(f"=== TEACHER VIEW SUBMISSION DEBUG ===")
    print(f"Submission ID: {submission_id}")
    print(f"Submission answers type: {type(submission.answers)}")
    print(f"Submission answers length: {len(submission.answers) if submission.answers else 0}")
    print(f"Submission questions count: {len(submission.questions) if submission.questions else 0}")
    
    # PROPER DATA PROCESSING: Use submission's stored questions and answers
    processed_answers = []
    
    if submission.questions and submission.answers:
        print(f"Processing {len(submission.questions)} questions from submission")
        
        for idx, question in enumerate(submission.questions):
            # Get question details
            if isinstance(question, dict):
                q_id = str(question.get('_id', question.get('id', '')))
                question_text = question.get('question_text', f'Question {idx + 1}')
                marks = question.get('marks', 2)
                model_answer = question.get('model_answer', '')
                question_type = question.get('question_type', 'subjective')
            else:
                q_id = str(getattr(question, '_id', getattr(question, 'id', '')))
                question_text = getattr(question, 'question_text', f'Question {idx + 1}')
                marks = getattr(question, 'marks', 2)
                model_answer = getattr(question, 'model_answer', '')
                question_type = getattr(question, 'question_type', 'subjective')
            
            # Get student answer
            student_answer = ''
            awarded_marks = 0
            is_correct = False
            
            if q_id in submission.answers:
                answer_data = submission.answers[q_id]
                if isinstance(answer_data, dict):
                    student_answer = answer_data.get('student_answer', '')
                    awarded_marks = answer_data.get('awarded_marks', 0)
                    is_correct = answer_data.get('is_correct', False)
                else:
                    student_answer = str(answer_data) if answer_data else ''
            
            print(f"Question {idx + 1} (ID: {q_id}): '{student_answer[:50]}...'")
            
            processed_answer = {
                'question_id': q_id,
                'question_text': question_text,
                'student_answer': student_answer,
                'marks': marks,
                'awarded_marks': awarded_marks,
                'is_correct': is_correct,
                'model_answer': model_answer,
                'question_type': question_type,
                'options': question.get('options', []) if isinstance(question, dict) else getattr(question, 'options', []),
                'correct_answer': question.get('correct_answer', '') if isinstance(question, dict) else getattr(question, 'correct_answer', '')
            }
            processed_answers.append(processed_answer)
    
    print(f"Final processed_answers count: {len(processed_answers)}")
    print(f"=== END TEACHER VIEW SUBMISSION DEBUG ===\n")
    
    # ALWAYS create processed answers from submission's questions (the ones student actually saw)
    # Use submission.questions if available, otherwise fall back to exam.questions
    questions_to_process = submission.questions if submission.questions else exam.questions
    if not processed_answers and questions_to_process:
        print("No processed answers found, creating fallback...")
        print(f"Processing {len(questions_to_process)} questions from submission")
        for idx, question in enumerate(questions_to_process):
            if isinstance(question, dict):
                q_id = str(question.get('_id', question.get('id', '')))
                question_text = question.get('question_text', f'Question {idx + 1}')
                marks = question.get('marks', 2)
                options = question.get('options', [])
                correct_answer = question.get('correct_answer', '')
                question_type = question.get('question_type', 'multiple_choice')
            else:
                q_id = str(getattr(question, '_id', getattr(question, 'id', '')))
                question_text = getattr(question, 'question_text', f'Question {idx + 1}')
                marks = getattr(question, 'marks', 2)
                options = getattr(question, 'options', [])
                correct_answer = getattr(question, 'correct_answer', '')
                question_type = getattr(question, 'question_type', 'multiple_choice')
            
            # Try to find student answer for this question
            student_answer = ''
            if submission.answers and isinstance(submission.answers, dict):
                # Try different possible keys
                possible_keys = [q_id, f'answer_{q_id}', str(idx), f'answer_{idx}']
                for key in possible_keys:
                    if key in submission.answers:
                        answer_data = submission.answers[key]
                        if isinstance(answer_data, dict):
                            # Extract from nested structure
                            student_answer = answer_data.get('student_answer', answer_data.get('answer', ''))
                            # If still a dict, try to extract further
                            if isinstance(student_answer, dict):
                                student_answer = student_answer.get('student_answer', student_answer.get('answer', ''))
                        else:
                            student_answer = str(answer_data) if answer_data else ''
                        break
            
            print(f"Question {idx + 1} (ID: {q_id}): Student answer = '{student_answer}'")
            
            processed_answer = {
                'question_id': q_id,
                'question_text': question_text,
                'student_answer': student_answer,
                'marks': marks,
                'options': options,
                'correct_answer': correct_answer,
                'model_answer': question.get('model_answer', '') if isinstance(question, dict) else getattr(question, 'model_answer', ''),
                'question_type': question_type
            }
            processed_answers.append(processed_answer)
            print(f"Created fallback answer for question {idx + 1}: {processed_answer}")
    
    print(f"=== END TEACHER VIEW SUBMISSION DEBUG ===\n")
    print(f"🔥🔥🔥 TEMPLATE RENDER DEBUG 🔥🔥🔥")
    print(f"SUBMISSION DATA:")
    print(f"  submission.score: {submission.score}")
    print(f"  submission.percentage: {submission.percentage}")
    print(f"  submission.max_score: {submission.max_score}")
    print(f"  submission.is_graded: {submission.is_graded}")
    print(f"PROCESSED ANSWERS:")
    print(f"  processed_answers type: {type(processed_answers)}")
    print(f"  processed_answers length: {len(processed_answers)}")
    if processed_answers:
        print(f"  First answer example: {processed_answers[0]}")
    print(f"🔥🔥🔥 END TEMPLATE RENDER DEBUG 🔥🔥🔥")
    
    return render_template('teacher/submission_details.html',
                         submission=submission,
                         exam=exam,
                         student=student,
                         processed_answers=processed_answers)

@teacher_bp.route('/submission/<submission_id>/grade', methods=['POST'])
@login_required
@teacher_required
def grade_submission(submission_id):
    # Get the submission
    submission = MongoManager.get_submission_by_id(submission_id)
    if not submission:
        flash('Submission not found', 'error')
        return redirect(url_for('teacher.dashboard'))
    
    # Get the exam
    exam = MongoManager.get_exam_by_id(submission.exam_id)
    if not exam:
        flash('Exam not found', 'error')
        return redirect(url_for('teacher.dashboard'))
    
    # Verify the exam belongs to the current teacher
    if str(exam.teacher_id) != str(current_user.id):
        flash('Unauthorized access', 'error')
        return redirect(url_for('teacher.dashboard'))
    
    try:
        # Get the overall score from the form (this is a percentage 0-100)
        score_percentage = request.form.get('score')
        if not score_percentage:
            flash('Please provide a total score', 'error')
            return redirect(url_for('teacher.view_submission', submission_id=submission_id))
        
        score_percentage = float(score_percentage)
        if score_percentage < 0 or score_percentage > 100:
            flash('Score must be between 0 and 100', 'error')
            return redirect(url_for('teacher.view_submission', submission_id=submission_id))
        
        # Calculate the actual marks from percentage
        # For subjective exams, we need to convert percentage to actual marks
        # Use submission's questions (what student actually saw) not all exam questions
        questions_for_calculation = submission.questions if submission.questions else exam.questions
        total_possible_marks = sum(q.get('marks', 2) if isinstance(q, dict) else getattr(q, 'marks', 2) for q in questions_for_calculation)
        actual_marks = (score_percentage / 100) * total_possible_marks
        
        print(f"🔥 GRADING DEBUG:")
        print(f"   Exam has {len(exam.questions)} total questions")
        print(f"   Submission has {len(questions_for_calculation)} questions (what student saw)")
        print(f"   Total possible marks: {total_possible_marks}")
        print(f"   Score percentage entered: {score_percentage}%")
        print(f"   Calculated actual marks: {actual_marks}")
        print(f"   This means: {actual_marks}/{total_possible_marks} = {score_percentage}%")
        
        print(f"Teacher grading: {score_percentage}% = {actual_marks}/{total_possible_marks} marks")
        
        # Get additional grading data for subjective exams
        pass_fail = request.form.get('pass_fail', '')
        feedback = request.form.get('feedback', '')
        
        # If this is a subjective exam, process question-wise scores
        if exam.exam_format == 'subjective':
            # PROPER GRADING: Update answers with grading information
            updated_answers = {}
            
            # Process each question from the submission
            for idx, question in enumerate(submission.questions):
                # Get question details
                if isinstance(question, dict):
                    q_id = str(question.get('_id', question.get('id', '')))
                    max_marks = question.get('marks', 2)
                else:
                    q_id = str(getattr(question, '_id', getattr(question, 'id', '')))
                    max_marks = getattr(question, 'marks', 2)
                
                # Get grading input from form
                question_score = request.form.get(f'question_{idx}_score', '0')
                question_status = request.form.get(f'question_{idx}_status', 'incorrect')
                
                try:
                    question_score = float(question_score)
                    max_marks = float(max_marks)
                    
                    # Ensure score doesn't exceed maximum marks
                    if question_score > max_marks:
                        question_score = max_marks
                    
                    # Get existing answer data or create new
                    if q_id in submission.answers:
                        answer_data = submission.answers[q_id].copy()
                    else:
                        answer_data = {
                            'student_answer': '',
                            'timestamp': datetime.utcnow().isoformat()
                        }
                    
                    # Update with grading information
                    answer_data.update({
                        'marks': max_marks,
                        'awarded_marks': question_score,
                        'is_correct': question_status == 'correct',
                        'status': question_status,
                        'graded_by': current_user.id,
                        'graded_at': datetime.utcnow().isoformat()
                    })
                    
                    updated_answers[q_id] = answer_data
                    
                except (ValueError, TypeError):
                    # If conversion fails, set to 0
                    answer_data = submission.answers.get(q_id, {}).copy()
                    answer_data.update({
                        'marks': max_marks,
                        'awarded_marks': 0,
                        'is_correct': False,
                        'status': 'incorrect'
                    })
                    updated_answers[q_id] = answer_data
        
        else:
            # For objective exams, convert percentage to actual marks and update
            # Use submission's questions (what student actually saw) not all exam questions
            questions_for_calculation = submission.questions if submission.questions else exam.questions
            total_possible_marks = sum(q.get('marks', 2) if isinstance(q, dict) else getattr(q, 'marks', 2) for q in questions_for_calculation)
            actual_marks = (score_percentage / 100) * total_possible_marks
            
            # Update with both actual marks and percentage
            from app.mongodb import mongo
            from bson import ObjectId
            
            update_data = {
                'score': actual_marks,
                'percentage': score_percentage,
                'max_score': total_possible_marks,
                'is_graded': True,
                'graded_at': datetime.utcnow(),
                'graded_by': current_user.id
            }
            
            result = mongo.db.submissions.update_one(
                {'_id': ObjectId(submission_id)},
                {'$set': update_data}
            )
            
            if result.modified_count > 0:
                flash('Submission graded successfully', 'success')
            else:
                flash('Failed to grade submission', 'error')
        
        # Update the submission in the database with the graded answers
        from app.mongodb import mongo
        from bson import ObjectId
        
        update_data = {
            'answers': updated_answers,
            'score': actual_marks,  # Store actual marks earned
            'percentage': score_percentage,  # Store the percentage separately
            'max_score': total_possible_marks,  # Store total possible marks
            'is_graded': True,
            'graded_at': datetime.utcnow(),
            'pass_fail': pass_fail,
            'feedback': feedback,
            'graded_by': current_user.id
        }
        
        result = mongo.db.submissions.update_one(
            {'_id': ObjectId(submission_id)},
            {'$set': update_data}
        )
        
        if result.modified_count > 0:
            flash('Submission graded successfully with detailed feedback!', 'success')
        else:
            flash('Failed to update submission with detailed grading', 'error')
        
    except (TypeError, ValueError) as e:
        print(f"Error grading submission: {str(e)}")
        flash('Invalid score value', 'error')
    except Exception as e:
        print(f"Unexpected error grading submission: {str(e)}")
        flash('An error occurred while grading the submission', 'error')
    
    return redirect(url_for('teacher.view_submission', submission_id=submission_id))

@teacher_bp.route('/exam/<exam_id>/update', methods=['POST'])
@login_required
@teacher_required
def update_exam(exam_id):
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     flash('Form validation failed. Please try again.', 'error')
    #     return redirect(url_for('teacher.exam_details', exam_id=exam_id))
            
    try:
        # Get form data
        title = request.form.get('title')
        description = request.form.get('description')
        duration = request.form.get('duration')
        total_marks = request.form.get('total_marks')
        exam_type = 'self_paced'  # Force all exams to be self_paced
        exam_format = request.form.get('exam_format', exam.exam_format)  # Keep existing format if not provided
        end_time = request.form.get('end_time')
        num_questions_to_display = request.form.get('num_questions_to_display')
        class_id = request.form.get('class_id', '')  # Get class_id, empty string if not provided
        is_active = request.form.get('is_active') == 'on'  # Convert checkbox value to boolean
        
        print(f"\n=== Updating Exam {exam_id} ===")
        print(f"Form data received:")
        print(f"title: {title}")
        print(f"description: {description}")
        print(f"duration: {duration}")
        print(f"total_marks: {total_marks}")
        print(f"exam_type: {exam_type}")
        print(f"exam_format: {exam_format}")
        print(f"end_time: {end_time}")
        print(f"num_questions_to_display: {num_questions_to_display}")
        print(f"class_id: {class_id}")
        print(f"is_active: {is_active}")
        
        # Validate required fields
        if not all([title, duration, total_marks, end_time]):
            return jsonify({'error': 'Please fill in all required fields'}), 400
            
        # Convert duration and total_marks to integers
        try:
            duration = int(duration)
            total_marks = int(total_marks)
            
            if duration < 1 or duration > 480:
                return jsonify({'error': 'Duration must be between 1 and 480 minutes'}), 400
                
            if total_marks < 1 or total_marks > 1000:
                return jsonify({'error': 'Total marks must be between 1 and 1000'}), 400
        
            # Convert num_questions_to_display to integer if provided
            if num_questions_to_display:
                num_questions_to_display = int(num_questions_to_display)
                if num_questions_to_display < 1:
                    return jsonify({'error': 'Number of questions to display must be at least 1'}), 400
            else:
                num_questions_to_display = None
                
        except ValueError as e:
            print(f"Error converting numeric values: {str(e)}")
            return jsonify({'error': 'Invalid numeric values provided'}), 400
            
        # Convert string dates to datetime
        try:
            # Handle end_time
            if not end_time:
                return jsonify({'error': 'Due date is required'}), 400
            
            # Parse end_time
            try:
                # First try direct ISO format
                end_time = datetime.fromisoformat(end_time)
            except ValueError:
                # If that fails, try parsing the date-time picker format
                end_time = datetime.strptime(end_time, '%Y-%m-%dT%H:%M')
            
            # Self-paced exams don't have start time
            start_time = None
            
            # Ensure end time is in the future
            if end_time <= datetime.utcnow():
                return jsonify({'error': 'Due date must be in the future'}), 400
                
            print(f"Parsed dates - end: {end_time}")
            
        except ValueError as e:
            print(f"Error parsing dates: {str(e)}")
            return jsonify({'error': 'Invalid date format. Please use format: YYYY-MM-DD HH:MM'}), 400
            
        # Convert class_id to ObjectId if provided
        if class_id:
            try:
                class_id = ObjectId(class_id)
                print(f"Converted class_id to ObjectId: {class_id}")
            except Exception as e:
                print(f"Error converting class_id: {str(e)}")
                return jsonify({'error': 'Invalid class ID format'}), 400
        else:
            class_id = None
            print("No class_id provided, setting to None")
        
        # Update the exam
        update_data = {
            'title': title,
            'description': description,
            'duration': duration,
            'total_marks': total_marks,
            'exam_type': exam_type,
            'exam_format': exam_format,
            'start_time': start_time,
            'end_time': end_time,
            'num_questions_to_display': num_questions_to_display,
            'class_id': class_id,
            'is_active': is_active
        }
        
        print(f"Update data prepared: {update_data}")
        
        if MongoManager.update_exam(exam_id, update_data):
            print("Exam updated successfully")
            return jsonify({'success': True, 'message': 'Exam updated successfully'})
        else:
            print("Failed to update exam in database")
            return jsonify({'error': 'Failed to update exam in database'}), 500
            
    except Exception as e:
        print(f"Error updating exam: {str(e)}")
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

@teacher_bp.route('/class/<class_id>/question-bank')
@login_required
@teacher_required
def class_question_bank(class_id):
    class_obj = MongoManager.get_class_by_id(class_id)
    if not class_obj or str(class_obj.teacher_id) != str(current_user.id):
        abort(404)
    
    questions = MongoManager.get_teacher_question_bank(current_user.id, class_id)
    return render_template('teacher/class_question_bank.html', 
                         class_obj=class_obj, 
                         questions=questions)

@teacher_bp.route('/class/<class_id>/question-bank/add', methods=['GET', 'POST'])
@login_required
@teacher_required
def add_to_class_bank(class_id):
    class_obj = MongoManager.get_class_by_id(class_id)
    if not class_obj or str(class_obj.teacher_id) != str(current_user.id):
        abort(404)
    
    if request.method == 'POST':
        question_data = {
            'text': request.form['question_text'],
            'type': request.form['question_type'],
            'points': int(request.form['points']),
            'difficulty_level': request.form['difficulty_level'],
            'subject': request.form['subject'],
            'topic': request.form['topic']
        }
        
        if question_data['type'] == 'multiple_choice':
            options = request.form.getlist('options[]')
            correct_answer = int(request.form['correct_answer'])
            question_data.update({
                'options': options,
                'correct_answer': correct_answer
            })
        else:
            question_data['model_answer'] = request.form.get('model_answer')
        
        if MongoManager.add_to_question_bank(question_data, current_user.id, class_id):
            flash('Question added to class bank successfully!', 'success')
        else:
            flash('Failed to add question to class bank.', 'error')
        return redirect(url_for('teacher.class_question_bank', class_id=class_id))
    
    return render_template('teacher/add_class_question.html', class_obj=class_obj)

@teacher_bp.route('/question-bank/view/<question_id>')
@login_required
@teacher_required
def view_bank_question(question_id):
    try:
        # Get the question from the bank
        questions = MongoManager.get_teacher_question_bank(current_user.id)
        question = next((q for q in questions if str(q['_id']) == question_id), None)
        
        if not question:
            return jsonify({'success': False, 'error': 'Question not found'})
        
        # Render the question details template
        html = render_template('teacher/_question_details.html', question=question)
        return jsonify({'success': True, 'html': html})
    except Exception as e:
        print(f"Error viewing question: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@teacher_bp.route('/question-bank/delete/<question_id>', methods=['POST'])
@login_required
@teacher_required
def delete_bank_question(question_id):
    try:
        if MongoManager.delete_from_question_bank(question_id, current_user.id):
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Failed to delete question'})
    except Exception as e:
        print(f"Error deleting question: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@teacher_bp.route('/exam/<exam_id>/questions/add', methods=['POST'])
@login_required
@teacher_required
def add_question(exam_id):
    # Skip CSRF validation since it's disabled in the app
    # from flask_wtf import FlaskForm
    # form = FlaskForm()
    
    # if not form.validate():
    #     flash('Form validation failed. Please try again.', 'error')
    #     return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
    
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam or exam.teacher_id != current_user.id:
        abort(404)
    
    try:
        # Get basic question data
        question_type = request.form.get('question_type')
        question_text = request.form.get('question_text')
        marks = int(request.form.get('marks', 1))
        
        # Validate required fields
        if not all([question_type, question_text]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        # Validate question type matches exam format
        if exam.is_objective() and question_type != 'multiple_choice':
            flash('Only multiple choice questions are allowed in objective exams.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        elif exam.is_subjective() and question_type != 'subjective':
            flash('Only subjective questions are allowed in subjective exams.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        # Prepare question data
        question_data = {
            'question_type': question_type,
            'question_text': question_text,
            'marks': marks
        }
        
        # Handle multiple choice questions
        if question_type == 'multiple_choice':
            options = request.form.getlist('options[]')
            correct_answer = request.form.get('correct_answer')
            
            # Validate multiple choice data
            if not options or not correct_answer:
                flash('Please provide all options and select a correct answer for multiple choice questions.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            # Clean and validate options
            options = [opt.strip() for opt in options if opt.strip()]
            if len(options) < 2:
                flash('Multiple choice questions must have at least 2 options.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            try:
                correct_answer_index = int(correct_answer)
                if not (0 <= correct_answer_index < len(options)):
                    flash('Invalid correct answer selection.', 'error')
                    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
                
                # Store the actual answer text instead of index
                correct_answer_text = options[correct_answer_index].strip()
                
            except ValueError:
                flash('Invalid correct answer value.', 'error')
                return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
            
            question_data.update({
                'options': options,
                'correct_answer': correct_answer_text  # Store actual text, not index
            })
        else:  # descriptive
            model_answer = request.form.get('model_answer', '').strip()
            question_data['model_answer'] = model_answer
        
        # Add the question to the exam
        if MongoManager.add_question(exam_id, question_data):
            flash('Question added successfully!', 'success')
        else:
            flash('Failed to add question. Please try again.', 'error')
            
    except Exception as e:
        print(f"Error adding question: {str(e)}")
        flash('An error occurred while adding the question.', 'error')
    
    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))

@teacher_bp.route('/exam/<exam_id>/import-questions', methods=['POST'])
@login_required
@teacher_required
def import_questions(exam_id):
    if 'questions_file' not in request.files:
        flash('No file uploaded.', 'error')
        return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
    
    exam = MongoManager.get_exam_by_id(exam_id)
    if not exam or exam.teacher_id != current_user.id:
        abort(404)
        
    file = request.files['questions_file']
    if not file.filename.endswith('.json'):
        flash('Only JSON files are supported.', 'error')
        return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
    
    try:
        data = json.loads(file.read())
        
        # Handle both formats: direct list or questions wrapper
        if isinstance(data, dict) and 'questions' in data:
            questions = data['questions']
        elif isinstance(data, list):
            questions = data
        else:
            flash('Invalid JSON format. Expected a list of questions or a dictionary with a "questions" key.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        if not isinstance(questions, list):
            flash('Invalid JSON format. Questions must be a list.', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        errors = []
        success_count = 0
        
        # Process each question
        for index, q in enumerate(questions, 1):
            try:
                # Validate required fields
                if not all(key in q for key in ['question_type', 'question_text', 'marks']):
                    errors.append(f'Question {index}: Missing required fields')
                    continue
                
                # Normalize question type to lowercase
                question_type = q['question_type'].lower()
                if question_type not in ['multiple_choice', 'subjective', 'descriptive']:
                    errors.append(f'Question {index}: Invalid question type "{q["question_type"]}" - must be "multiple_choice" or "subjective"')
                    continue
                
                # Convert 'descriptive' to 'subjective' for consistency
                if question_type == 'descriptive':
                    question_type = 'subjective'
                
                # Validate question type matches exam format
                if exam.is_objective() and question_type != 'multiple_choice':
                    errors.append(f'Question {index}: Only multiple choice questions are allowed in objective exams')
                    continue
                elif exam.is_subjective() and question_type != 'subjective':
                    errors.append(f'Question {index}: Only subjective questions are allowed in subjective exams')
                    continue
                
                # Prepare question data
                question_data = {
                    'question_type': question_type,
                    'question_text': q['question_text'].strip(),
                    'marks': int(q['marks'])
                }
                
                # Validate marks
                if question_data['marks'] < 1:
                    errors.append(f'Question {index}: Marks must be greater than 0')
                    continue
                
                # Handle multiple choice questions
                if question_type == 'multiple_choice':
                    if 'options' not in q or 'correct_answer' not in q:
                        errors.append(f'Question {index}: Multiple choice questions must have options and correct_answer')
                        continue
                    
                    options = q['options']
                    if not isinstance(options, list) or len(options) < 2:
                        errors.append(f'Question {index}: Multiple choice questions must have at least 2 options')
                        continue
                    
                    # Clean options
                    options = [str(opt).strip() for opt in options if str(opt).strip()]
                    if len(options) < 2:
                        errors.append(f'Question {index}: Multiple choice questions must have at least 2 non-empty options')
                        continue
                    
                    try:
                        correct_answer_index = int(q['correct_answer'])
                        if not (0 <= correct_answer_index < len(options)):
                            errors.append(f'Question {index}: Correct answer index must be between 0 and {len(options)-1}')
                            continue
                        
                        # Store the actual answer text instead of index
                        correct_answer_text = options[correct_answer_index].strip()
                        
                    except (ValueError, TypeError):
                        errors.append(f'Question {index}: Invalid correct answer value')
                        continue
                    
                    question_data.update({
                        'options': options,
                        'correct_answer': correct_answer_text  # Store actual text, not index
                    })
                else:  # subjective
                    question_data['model_answer'] = q.get('model_answer', '').strip()
                
                # Add the question to the exam
                if MongoManager.add_question(exam_id, question_data):
                    success_count += 1
                else:
                    errors.append(f'Question {index}: Failed to add to database')
                
            except Exception as e:
                errors.append(f'Question {index}: {str(e)}')
                continue
        
        # Report results
        if success_count > 0:
            flash(f'Successfully imported {success_count} questions.', 'success')
        if errors:
            flash('Some questions could not be imported:\n' + '\n'.join(errors), 'warning')
            
    except json.JSONDecodeError:
        flash('Invalid JSON file format.', 'error')
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'error')
        
    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
@teacher_bp.route('/exam/<exam_id>/questions/delete-all', methods=['POST'])
@login_required
@teacher_required
def delete_all_questions(exam_id):
    try:
        # Get the exam
        exam = MongoManager.get_exam_by_id(exam_id)
        if not exam or str(exam.teacher_id) != str(current_user.id):
            flash('Unauthorized access', 'error')
            return redirect(url_for('teacher.manage_questions', exam_id=exam_id))
        
        # Delete all questions
        if MongoManager.delete_all_questions(exam_id):
            flash('All questions have been deleted successfully!', 'success')
        else:
            flash('Failed to delete questions.', 'error')
            
    except Exception as e:
        print(f"Error deleting all questions: {str(e)}")
        flash('An error occurred while deleting questions.', 'error')
    
    return redirect(url_for('teacher.manage_questions', exam_id=exam_id))

@teacher_bp.route('/exam/<exam_id>/export-excel', methods=['GET'])
@login_required
@teacher_required
def export_excel(exam_id):
    """Export comprehensive exam data to Excel with student submissions focus"""
    try:
        exam = MongoManager.get_exam_by_id(exam_id)
        if not exam or str(exam.teacher_id) != str(current_user.id):
            abort(404)
        
        # Get submissions for this exam
        submissions = MongoManager.get_exam_submissions(exam_id)
        print(f"Found {len(submissions)} submissions for exam {exam_id}")
        
        # Get all students to show comprehensive list
        all_students = MongoManager.get_all_students()
        print(f"Found {len(all_students)} total students")
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        title_font = Font(bold=True, size=16)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ===== Main Sheet: Student Submissions =====
        ws = wb.create_sheet(f"{exam.title} - Submissions")
        
        # Title and exam info
        ws.append([f"ExamPro - {exam.title} Submissions"])
        ws["A1"].font = title_font
        ws.append([])
        ws.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"Teacher: {current_user.username}"])
        ws.append([f"Exam: {exam.title}"])
        ws.append([f"Total Submissions: {len(submissions)}"])
        ws.append([f"Total Students: {len(all_students)}"])
        ws.append([])
        
        # Headers - comprehensive student data
        headers = ["#", "Student Name", "Email", "Student ID", "Phone", "College", 
                  "Submitted At", "Score", "Status", "Graded By", "Graded At", 
                  "Pass/Fail", "Time Taken", "Violations", "Auto-Submit"]
        ws.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Create a dict of submitted students for quick lookup
        submitted_students = {}
        for submission in submissions:
            if hasattr(submission, 'student_id'):
                submitted_students[str(submission.student_id)] = submission
        
        row_num = 10
        student_count = 1
        
        # Process all students (submitted and not submitted)
        for student in all_students:
            try:
                student_id = str(student.id)
                submission = submitted_students.get(student_id)
                
                if submission:
                    # Student has submitted
                    # Format submission date
                    submitted_at = 'N/A'
                    if hasattr(submission, 'submitted_at') and submission.submitted_at:
                        try:
                            submitted_at = submission.submitted_at.strftime('%Y-%m-%d %H:%M')
                        except:
                            submitted_at = str(submission.submitted_at)
                    
                    # Format score
                    score_str = 'Pending'
                    if hasattr(submission, 'score') and submission.score is not None:
                        try:
                            score_str = f"{float(submission.score):.1f}%"
                        except:
                            score_str = str(submission.score)
                    
                    # Determine status
                    status = "Pending"
                    if getattr(submission, 'is_graded', False):
                        status = "Graded"
                    
                    # Get grader info
                    grader_name = "N/A"
                    if hasattr(submission, 'graded_by') and submission.graded_by:
                        try:
                            grader = MongoManager.get_user_by_id(submission.graded_by)
                            grader_name = getattr(grader, 'username', 'Unknown') if grader else "Unknown"
                        except:
                            grader_name = "Unknown"
                    elif status == "Graded" and getattr(exam, 'exam_format', '') == 'objective':
                        grader_name = "Auto-graded"
                    
                    # Format graded date
                    graded_at = 'N/A'
                    if hasattr(submission, 'graded_at') and submission.graded_at:
                        try:
                            graded_at = submission.graded_at.strftime('%Y-%m-%d %H:%M')
                        except:
                            graded_at = str(submission.graded_at)
                    elif status == "Graded":
                        graded_at = submitted_at
                    
                    # Pass/Fail
                    pass_fail = "N/A"
                    if hasattr(submission, 'score') and submission.score is not None:
                        try:
                            pass_fail = "Pass" if float(submission.score) >= 30 else "Fail"
                        except:
                            pass_fail = "N/A"
                    
                    # Time taken
                    time_taken = "N/A"
                    if hasattr(submission, 'started_at') and hasattr(submission, 'submitted_at'):
                        if submission.started_at and submission.submitted_at:
                            try:
                                time_diff = submission.submitted_at - submission.started_at
                                time_taken = f"{time_diff.total_seconds() / 60:.1f} min"
                            except:
                                pass
                    
                    # Violations and auto-submit
                    violations = getattr(submission, 'warning_count_at_submission', 0) or 0
                    auto_submit = "Yes" if getattr(submission, 'auto_submitted', False) else "No"
                    
                else:
                    # Student has not submitted
                    submitted_at = "Not submitted"
                    score_str = "Not attempted"
                    status = "Not submitted"
                    grader_name = "N/A"
                    graded_at = "N/A"
                    pass_fail = "N/A"
                    time_taken = "N/A"
                    violations = 0
                    auto_submit = "No"
                
                # Prepare row data
                row_data = [
                    student_count,
                    getattr(student, 'username', 'Unknown'),
                    getattr(student, 'email', 'N/A'),
                    getattr(student, 'student_id', 'N/A'),
                    getattr(student, 'phone_number', 'N/A'),
                    getattr(student, 'college_name', 'N/A'),
                    submitted_at,
                    score_str,
                    status,
                    grader_name,
                    graded_at,
                    pass_fail,
                    time_taken,
                    violations,
                    auto_submit
                ]
                
                ws.append(row_data)
                
                # Style data rows with conditional formatting
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                    
                    # Color coding
                    if not submission:  # Not submitted
                        if col_num > 1:  # Don't color the index
                            cell.fill = PatternFill("solid", fgColor="F0F0F0")  # Light gray
                    else:
                        # Color coding for submitted students
                        if col_num == 12:  # Pass/Fail column
                            if pass_fail == "Pass":
                                cell.fill = PatternFill("solid", fgColor="90EE90")  # Light green
                            elif pass_fail == "Fail":
                                cell.fill = PatternFill("solid", fgColor="FFB6C1")  # Light red
                        
                        if col_num == 15 and auto_submit == "Yes":  # Auto-submit column
                            cell.fill = PatternFill("solid", fgColor="FFA500")  # Orange
                
                row_num += 1
                student_count += 1
                
            except Exception as e:
                print(f"Error processing student {getattr(student, 'username', 'Unknown')}: {str(e)}")
                continue
        
        # ===== Summary Sheet =====
        ws_summary = wb.create_sheet("Summary")
        
        # Calculate statistics
        total_students = len(all_students)
        total_submitted = len(submissions)
        not_submitted = total_students - total_submitted
        graded_submissions = [sub for sub in submissions if getattr(sub, 'is_graded', False)]
        
        # Calculate scores
        scores = [sub.score for sub in graded_submissions if hasattr(sub, 'score') and sub.score is not None]
        avg_score = sum(scores) / len(scores) if scores else 0
        highest_score = max(scores) if scores else 0
        lowest_score = min(scores) if scores else 0
        
        # Pass/fail counts
        passed = len([score for score in scores if score >= 30])
        failed = len([score for score in scores if score < 30])
        
        # Auto-submit stats
        auto_submitted = len([sub for sub in submissions if getattr(sub, 'auto_submitted', False)])
        
        # Summary data
        ws_summary.append([f"Exam Summary: {exam.title}"])
        ws_summary["A1"].font = title_font
        ws_summary.append([])
        ws_summary.append(["Exam Information", ""])
        ws_summary.append(["Title", exam.title])
        ws_summary.append(["Duration", f"{exam.duration} minutes"])
        ws_summary.append(["Total Marks", getattr(exam, 'total_marks', 'N/A')])
        ws_summary.append(["Format", getattr(exam, 'exam_format', 'N/A').title()])
        ws_summary.append([])
        ws_summary.append(["Participation Statistics", ""])
        ws_summary.append(["Total Students", total_students])
        ws_summary.append(["Students Submitted", total_submitted])
        ws_summary.append(["Students Not Submitted", not_submitted])
        ws_summary.append(["Participation Rate", f"{(total_submitted/total_students*100):.1f}%" if total_students > 0 else "0%"])
        ws_summary.append([])
        ws_summary.append(["Performance Statistics", ""])
        ws_summary.append(["Graded Submissions", len(graded_submissions)])
        ws_summary.append(["Average Score", f"{avg_score:.1f}%" if avg_score > 0 else "0%"])
        ws_summary.append(["Highest Score", f"{highest_score:.1f}%" if highest_score > 0 else "0%"])
        ws_summary.append(["Lowest Score", f"{lowest_score:.1f}%" if lowest_score > 0 else "0%"])
        ws_summary.append(["Students Passed", passed])
        ws_summary.append(["Students Failed", failed])
        ws_summary.append(["Pass Rate", f"{(passed/(passed+failed)*100):.1f}%" if (passed+failed) > 0 else "0%"])
        ws_summary.append(["Auto-Submitted", auto_submitted])
        
        # Auto-adjust column widths
        try:
            for sheet in wb.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as width_error:
            print(f"Error adjusting column widths: {str(width_error)}")
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        try:
            exam_title = getattr(exam, 'title', 'Exam')
            safe_title = "".join(c for c in exam_title if c.isalnum() or c in (' ', '-', '_')).rstrip()
            if not safe_title:
                safe_title = "Exam"
            filename = f"{safe_title}_Student_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        except:
            filename = f"Exam_Student_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error generating Excel report. Please try again.', 'error')
        return redirect(url_for('teacher.exam_details', exam_id=exam_id))

@teacher_bp.route('/students/export-excel', methods=['GET'])
@login_required
@teacher_required
def export_students_excel():
    """Export comprehensive student records to Excel"""
    try:
        # Get all students
        all_students = MongoManager.get_all_students()
        
        # Get teacher's classes to filter relevant students
        teacher_classes = MongoManager.get_teacher_classes(current_user.id)
        teacher_exam_ids = [exam.id for exam in MongoManager.find_exams(teacher_id=current_user.id)]
        
        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ===== Sheet 1: Student Overview =====
        ws_overview = wb.create_sheet("Student Overview")
        
        # Title
        ws_overview.append(["ExamPro - Student Records Report"])
        ws_overview["A1"].font = Font(bold=True, size=16)
        ws_overview.append([])
        ws_overview.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_overview.append([f"Teacher: {current_user.username}"])
        ws_overview.append([])
        
        # Headers for student overview
        overview_headers = ["#", "Student Name", "Email", "Student ID", "Class", "Total Exams Taken", "Average Score (%)", "Highest Score (%)", "Lowest Score (%)", "Status"]
        ws_overview.append(overview_headers)
        
        # Style headers
        for col_num, header in enumerate(overview_headers, 1):
            cell = ws_overview.cell(row=6, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Process each student
        row_num = 7
        for idx, student in enumerate(all_students, 1):
            try:
                # Get student's exam submissions
                student_submissions = MongoManager.get_student_submissions(student.id)
                teacher_submissions = [sub for sub in student_submissions if str(sub.exam_id) in [str(eid) for eid in teacher_exam_ids]]
                
                # Calculate statistics
                total_exams = len(teacher_submissions)
                scores = [sub.score for sub in teacher_submissions if hasattr(sub, 'score') and sub.score is not None]
                avg_score = sum(scores) / len(scores) if scores else 0
                highest_score = max(scores) if scores else 0
                lowest_score = min(scores) if scores else 0
                
                # Determine status
                status = "Active" if getattr(student, 'is_active', True) else "Inactive"
                
                # Get student's class
                student_class = "Not assigned"
                for class_obj in teacher_classes:
                    if hasattr(class_obj, 'student_ids') and student.id in class_obj.student_ids:
                        student_class = class_obj.name
                        break
                
                row_data = [
                    idx,
                    getattr(student, 'username', 'Unknown'),
                    getattr(student, 'email', 'N/A'),
                    getattr(student, 'student_id', 'N/A'),
                    student_class,
                    total_exams,
                    f"{avg_score:.1f}" if avg_score > 0 else "0.0",
                    f"{highest_score:.1f}" if highest_score > 0 else "0.0",
                    f"{lowest_score:.1f}" if lowest_score > 0 else "0.0",
                    status
                ]
                
                ws_overview.append(row_data)
                
                # Style rows
                for col_num in range(1, len(overview_headers) + 1):
                    cell = ws_overview.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                
                row_num += 1
                
            except Exception as e:
                print(f"Error processing student {getattr(student, 'username', 'Unknown')}: {str(e)}")
                continue
        
        # Auto-adjust column widths
        try:
            for sheet in wb.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as width_error:
            print(f"Error adjusting column widths: {str(width_error)}")
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"Students_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting students Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error generating students report. Please try again.', 'error')
        return redirect(url_for('teacher.students'))


@teacher_bp.route('/exam/<exam_id>/export-submissions', methods=['GET'])
@login_required
@teacher_required
def export_exam_submissions_excel(exam_id):
    """Export submissions for a specific exam to Excel"""
    try:
        # Verify exam exists and belongs to current teacher
        exam = MongoManager.get_exam_by_id(exam_id)
        if not exam:
            flash('Exam not found', 'error')
            return redirect(url_for('teacher.dashboard'))
        
        if str(exam.teacher_id) != str(current_user.id):
            flash('Unauthorized access', 'error')
            return redirect(url_for('teacher.dashboard'))
        
        # Get all submissions for this specific exam
        all_submissions = MongoManager.get_exam_submissions(exam_id)
        
        # Sort by submitted_at (most recent first)
        all_submissions.sort(key=lambda x: x.submitted_at if x.submitted_at else datetime.min, reverse=True)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ===== Sheet 1: Recent Submissions =====
        ws_submissions = wb.create_sheet("Recent Submissions")
        
        # Title
        ws_submissions.append([f"ExamPro - {exam.title} Submissions"])
        ws_submissions["A1"].font = Font(bold=True, size=16)
        ws_submissions.append([])
        ws_submissions.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_submissions.append([f"Teacher: {current_user.username}"])
        ws_submissions.append([f"Exam: {exam.title}"])
        ws_submissions.append([f"Total Submissions: {len(all_submissions)}"])
        ws_submissions.append([])
        
        # Headers matching the Recent Submissions table
        headers = ["Student", "Submitted At", "Score", "Status", "Graded By", "Graded At"]
        ws_submissions.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws_submissions.cell(row=8, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Process submissions data
        row_num = 9
        for submission in all_submissions:
            try:
                # Get student info
                student = MongoManager.get_user_by_id(submission.student_id)
                student_name = getattr(student, 'username', 'Unknown') if student else 'Unknown'
                
                # Format submission date
                submitted_at = 'N/A'
                if hasattr(submission, 'submitted_at') and submission.submitted_at:
                    try:
                        submitted_at = submission.submitted_at.strftime('%Y-%m-%d %H:%M')
                    except:
                        submitted_at = str(submission.submitted_at)
                
                # Format score
                score_str = 'Pending'
                if hasattr(submission, 'score') and submission.score is not None:
                    try:
                        score_str = f"{float(submission.score):.1f}%"
                    except:
                        score_str = str(submission.score)
                
                # Determine status
                status = "Pending"
                if getattr(submission, 'is_graded', False):
                    status = "Graded"
                
                # Get grader info
                grader_name = "N/A"
                if hasattr(submission, 'graded_by') and submission.graded_by:
                    try:
                        grader = MongoManager.get_user_by_id(submission.graded_by)
                        grader_name = getattr(grader, 'username', 'Unknown') if grader else "Unknown"
                    except:
                        grader_name = "Unknown"
                elif status == "Graded":
                    grader_name = "System"
                
                # Format graded date
                graded_at = 'N/A'
                if hasattr(submission, 'graded_at') and submission.graded_at:
                    try:
                        graded_at = submission.graded_at.strftime('%Y-%m-%d %H:%M')
                    except:
                        graded_at = str(submission.graded_at)
                elif status == "Graded" and graded_at == 'N/A':
                    graded_at = submitted_at  # Use submission date if graded_at is not available
                
                row_data = [
                    student_name,
                    submitted_at,
                    score_str,
                    status,
                    grader_name,
                    graded_at
                ]
                
                ws_submissions.append(row_data)
                
                # Style data rows
                for col_num in range(1, len(headers) + 1):
                    cell = ws_submissions.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                
                row_num += 1
                
            except Exception as submission_error:
                print(f"Error processing submission: {str(submission_error)}")
                continue
        
        # Auto-adjust column widths
        try:
            for sheet in wb.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as width_error:
            print(f"Error adjusting column widths: {str(width_error)}")
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        try:
            exam_title = getattr(exam, 'title', 'Exam')
            safe_title = "".join(c for c in exam_title if c.isalnum() or c in (' ', '-', '_')).rstrip()
            if not safe_title:
                safe_title = "Exam"
            filename = f"{safe_title}_Submissions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        except:
            filename = f"Exam_Submissions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting submissions: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error generating submissions report. Please try again.', 'error')
        return redirect(url_for('teacher.exam_details', exam_id=exam_id))
